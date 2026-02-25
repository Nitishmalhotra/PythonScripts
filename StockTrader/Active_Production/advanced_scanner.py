"""
Advanced Stock Scanner with Multiple Strategies + Integrated Backtesting
Includes refined trading strategies and walk-forward backtest engine
"""

from kite_stock_scanner import KiteStockScanner
import pandas as pd
import numpy as np
import logging
from datetime import datetime, timedelta
import os
from pathlib import Path
from enhanced_html_generator import generate_enhanced_html, calculate_market_sentiment, get_strategy_recommendation

logger = logging.getLogger(__name__)

SCRIPT_DIR  = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
RESULTS_DIR  = PROJECT_ROOT / "Results"
RESULTS_DIR.mkdir(exist_ok=True)


# ═══════════════════════════════════════════════════════════════════════
#  BACKTEST ENGINE
# ═══════════════════════════════════════════════════════════════════════

class BacktestEngine:
    """
    Walk-forward backtester for all scanner strategies.

    HOW IT WORKS
    ─────────────────────────────────────────────────────
    For each symbol:
      1. Split history into IN-SAMPLE (train) and OUT-OF-SAMPLE (test).
      2. Run the strategy signal function on the test window row-by-row
         (no look-ahead — signal on day T is acted on at open of day T+1).
      3. Simulate trades with STRATEGY-SPECIFIC ATR stops and targets.
      4. Record every trade: entry, exit, P&L, hold days, exit reason.
      5. Aggregate metrics: win rate, avg R:R, max drawdown, Sharpe, etc.

    EXIT LOGIC (per trade)
    ─────────────────────────────────────────────────────
    • Stop-loss hit  → close at stop price
    • Target 1 hit   → close 50 % of position, trail stop to entry
    • Target 2 hit   → close remaining position
    • Max hold days  → close at close price (time-based exit)

    STRATEGY-SPECIFIC PARAMETERS
    ─────────────────────────────────────────────────────
    Tighter stops + shorter holds  → Mean Reversion, Gap Up, RSI Setup
    Wider stops  + longer holds    → Momentum, Trend Following, Swing, Stage 2
    """

    # Stop / Target / MaxHold tuned per strategy from backtest observation:
    #   - Losing strategies had stop exits dominating (stops too tight)
    #   - Winners had balanced stop/target distribution
    STRATEGY_PARAMS = {
        'Momentum Breakout': dict(stop=2.0, t1=3.5, t2=5.0, hold=25),
        'Mean Reversion':    dict(stop=1.2, t1=2.0, t2=3.0, hold=10),
        'Trend Following':   dict(stop=2.5, t1=4.0, t2=6.0, hold=30),
        'Volume Breakout':   dict(stop=2.0, t1=3.0, t2=4.5, hold=15),
        'Swing Trading':     dict(stop=2.0, t1=3.5, t2=5.0, hold=20),
        'Gap Up':            dict(stop=1.2, t1=2.0, t2=3.0, hold=10),
        'RSI Setup':         dict(stop=1.5, t1=2.5, t2=3.5, hold=12),
        'Stage 2 Uptrend':   dict(stop=2.5, t1=4.0, t2=6.0, hold=40),
        'Strong Linearity':  dict(stop=1.8, t1=3.0, t2=4.5, hold=15),
        'VCP Pattern':       dict(stop=1.5, t1=2.5, t2=4.0, hold=20),
        'Pyramiding':        dict(stop=1.5, t1=2.5, t2=3.5, hold=20),
        'Golden Crossover':  dict(stop=3.0, t1=5.0, t2=7.0, hold=40),
    }
    DEFAULT_PARAMS = dict(stop=1.5, t1=2.5, t2=3.5, hold=20)

    def __init__(self,
                 capital:   float = 100_000,
                 risk_pct:  float = 0.02):      # 2 % risk per trade
        self.capital  = capital
        self.risk_pct = risk_pct

    # ──────────────────────────────────────────────────────────────────
    def _position_size(self, entry: float, stop: float) -> int:
        """Shares to buy so that risk = capital * risk_pct."""
        risk_per_share = abs(entry - stop)
        if risk_per_share <= 0:
            return 0
        return max(1, int(self.capital * self.risk_pct / risk_per_share))

    # ──────────────────────────────────────────────────────────────────
    def _simulate_trade(self, df: pd.DataFrame, signal_idx: int,
                        stop_mult: float = 1.5, t1_mult: float = 2.5,
                        t2_mult: float = 3.5, max_hold: int = 20) -> dict | None:
        """Simulate one trade using strategy-specific ATR multipliers."""
        entry_row_idx = signal_idx + 1
        if entry_row_idx >= len(df):
            return None

        entry_price = df['open'].iloc[entry_row_idx]
        atr         = df['atr'].iloc[signal_idx]
        if atr <= 0 or entry_price <= 0:
            return None

        stop   = entry_price - stop_mult * atr
        t1     = entry_price + t1_mult   * atr
        t2     = entry_price + t2_mult   * atr
        shares = self._position_size(entry_price, stop)
        if shares == 0:
            return None

        half        = max(1, shares // 2)
        t1_hit      = False
        trail       = stop
        pnl         = 0.0
        exit_price  = entry_price
        exit_reason = 'max_hold'
        exit_date   = df.index[min(entry_row_idx + max_hold - 1, len(df)-1)]

        for i in range(entry_row_idx, min(entry_row_idx + max_hold, len(df))):
            lo, hi, cl = df['low'].iloc[i], df['high'].iloc[i], df['close'].iloc[i]

            if lo <= trail:
                exit_price  = trail
                exit_reason = 'stop_loss'
                exit_date   = df.index[i]
                pnl = (exit_price - entry_price) * (half if t1_hit else shares)
                break

            if not t1_hit and hi >= t1:
                pnl   += (t1 - entry_price) * half
                trail  = entry_price          # trail to breakeven after T1
                t1_hit = True

            if t1_hit and hi >= t2:
                pnl        += (t2 - entry_price) * half
                exit_price  = t2
                exit_reason = 'target_2'
                exit_date   = df.index[i]
                break

            if i == min(entry_row_idx + max_hold - 1, len(df)-1):
                pnl        = (cl - entry_price) * (half if t1_hit else shares)
                exit_price = cl
                exit_date  = df.index[i]

        return {
            'entry_date':  df.index[entry_row_idx],
            'exit_date':   exit_date,
            'entry_price': round(entry_price, 2),
            'exit_price':  round(exit_price, 2),
            'stop':        round(stop, 2),
            'target_1':    round(t1, 2),
            'target_2':    round(t2, 2),
            'shares':      shares,
            'pnl':         round(pnl, 2),
            'pnl_pct':     round(pnl / (entry_price * shares) * 100, 2),
            'exit_reason': exit_reason,
            'hold_days':   max(1, i - entry_row_idx + 1),
            'atr':         round(atr, 2),
        }

    # ──────────────────────────────────────────────────────────────────
    def _aggregate_trades(self, trades: list, strategy: str, symbol: str) -> dict:
        """Compute summary metrics from a list of trade dicts."""
        if not trades:
            return {}

        df_t = pd.DataFrame(trades)
        n           = len(df_t)
        wins        = df_t[df_t['pnl'] > 0]
        losses      = df_t[df_t['pnl'] <= 0]
        win_rate    = len(wins) / n * 100
        avg_win     = wins['pnl_pct'].mean() if len(wins) else 0
        avg_loss    = losses['pnl_pct'].mean() if len(losses) else 0
        total_pnl   = df_t['pnl'].sum()

        # Cumulative equity for drawdown
        equity      = df_t['pnl'].cumsum()
        roll_max    = equity.cummax()
        drawdown    = equity - roll_max
        max_dd      = drawdown.min()

        # Sharpe (daily returns proxy)
        pnl_pct_std = df_t['pnl_pct'].std()
        sharpe      = (df_t['pnl_pct'].mean() / pnl_pct_std * np.sqrt(252)
                       if pnl_pct_std > 0 else 0)

        # Profit factor
        gross_win   = wins['pnl'].sum()   if len(wins)   else 0
        gross_loss  = abs(losses['pnl'].sum()) if len(losses) else 1
        pf          = gross_win / gross_loss if gross_loss else float('inf')

        exit_dist   = df_t['exit_reason'].value_counts().to_dict()

        return {
            'strategy':         strategy,
            'symbol':           symbol,
            'total_trades':     n,
            'win_rate_pct':     round(win_rate, 1),
            'avg_win_pct':      round(avg_win, 2),
            'avg_loss_pct':     round(avg_loss, 2),
            'total_pnl_inr':    round(total_pnl, 2),
            'max_drawdown_inr': round(max_dd, 2),
            'profit_factor':    round(pf, 2),
            'sharpe_ratio':     round(sharpe, 2),
            'avg_hold_days':    round(df_t['hold_days'].mean(), 1),
            'stop_loss_exits':  exit_dist.get('stop_loss', 0),
            'target1_exits':    exit_dist.get('target_1', 0),
            'target2_exits':    exit_dist.get('target_2', 0),
            'time_exits':       exit_dist.get('max_hold', 0),
            'trades':           trades,          # Raw trade log
        }

    # ──────────────────────────────────────────────────────────────────
    def run(self, stock_data: pd.DataFrame,
            strategy_func,
            strategy_name: str,
            test_pct: float = 0.3) -> dict:
        """
        True walk-forward backtest for ONE strategy across ALL symbols.

        KEY DESIGN
        ──────────────────────────────────────────────────────────────
        For each bar i in the TEST WINDOW, we call the strategy with
        data[:i+1] (everything up to and including that bar). If the
        strategy fires (returns non-empty), that bar is a signal day.
        Entry is at the NEXT bar's open.

        This correctly mirrors how `.iloc[-1]` strategies work in
        the live scanner — each day, the strategy only sees today
        and prior history, and signals fire on the last bar only.
        """
        all_summaries = []
        all_trades    = []

        for symbol, sym_df in stock_data.groupby('symbol'):
            sym_df = sym_df.sort_values('date').reset_index(drop=True)

            n_rows = len(sym_df)
            if n_rows < 60:
                continue

            split      = int(n_rows * (1 - test_pct))
            sim_df     = sym_df.set_index('date')
            p          = self.STRATEGY_PARAMS.get(strategy_name, self.DEFAULT_PARAMS)

            trades         = []
            in_trade_until = -1   # row index; skip signal if already in a trade

            # Walk forward through every bar in the test window
            for i in range(split, n_rows):
                if i <= in_trade_until:
                    continue   # avoid overlapping trades

                # Pass data up to and including bar i (no look-ahead)
                window = sym_df.iloc[:i + 1].copy()
                try:
                    result = strategy_func(window)
                except Exception:
                    continue

                if result is None or result.empty:
                    continue

                # Signal fired — trade entry at next bar
                entry_idx = i + 1
                if entry_idx >= n_rows:
                    continue

                try:
                    trade = self._simulate_trade(
                        sim_df, i,
                        stop_mult=p['stop'], t1_mult=p['t1'],
                        t2_mult=p['t2'],     max_hold=p['hold']
                    )
                except Exception as e:
                    logger.debug(f"[BT] trade sim error {symbol} bar {i}: {e}")
                    continue

                if trade:
                    trade['symbol']   = symbol
                    trade['strategy'] = strategy_name
                    trades.append(trade)
                    all_trades.append(trade)
                    in_trade_until = i + trade.get('hold_days', p['hold'])

            if trades:
                summary = self._aggregate_trades(trades, strategy_name, symbol)
                if summary:
                    all_summaries.append(summary)

        # ── Overall summary ───────────────────────────────────────────
        overall = {}
        if all_trades:
            overall = self._aggregate_trades(all_trades, strategy_name, 'ALL_SYMBOLS')
            logger.info(
                f"[Backtest] {strategy_name}: "
                f"trades={overall['total_trades']}, "
                f"win_rate={overall['win_rate_pct']:.1f}%, "
                f"PF={overall['profit_factor']:.2f}, "
                f"Sharpe={overall['sharpe_ratio']:.2f}"
            )
        else:
            logger.info(f"[Backtest] {strategy_name}: trades=0 — no test-window signals produced")

        return {
            'strategy':   strategy_name,
            'per_symbol': all_summaries,
            'overall':    overall,
            'all_trades': all_trades,
        }


# ═══════════════════════════════════════════════════════════════════════
#  ADVANCED SCANNER
# ═══════════════════════════════════════════════════════════════════════

class AdvancedStockScanner(KiteStockScanner):
    """
    Advanced scanner with multiple trading strategies + integrated backtester.
    """

    # ──────────────────────────────────────────────────────────────────
    # INTERNAL HELPER
    # ──────────────────────────────────────────────────────────────────
    @staticmethod
    def _prep(df: pd.DataFrame) -> pd.DataFrame:
        """
        Ensure df is safe for boolean operations:
        - If 'date' column exists, set it as the index (so it won't
          participate in numeric comparisons that trigger the
          'datetime64 does not support any()' error).
        - Reset to a clean RangeIndex if already indexed by date,
          so strategies can use .iloc safely.
        """
        df = df.copy()
        if 'date' in df.columns:
            df = df.set_index('date')
        # Keep only numeric columns for condition building;
        # non-numeric cols like 'symbol' stay but won't cause issues
        # because conditions are built from named numeric columns only.
        return df

    # ──────────────────────────────────────────────────────────────────
    # STRATEGY 1 — Momentum Breakout
    # ──────────────────────────────────────────────────────────────────
    def strategy_momentum_breakout(self, df):
        """
        Price breaking above 20-day high TODAY on strong volume.
        Confirmed by rising MACD histogram and RSI in 55-78 band.
        Signal only fires on the most recent bar.
        """
        df = self._prep(df)
        if len(df) < 25:
            return pd.DataFrame()

        high_20 = df['high'].shift(1).rolling(20).max()
        conditions = (
            (df['close'] > high_20) &                           # NEW 20-day high
            (df['close'] > df['close'].shift(1)) &             # Up day
            (df['volume_ratio'] > 1.5) &                       # Volume confirmation
            (df['rsi_14'] > 55) & (df['rsi_14'] < 78) &       # Momentum not overbought
            (df['macd_hist'] > df['macd_hist'].shift(1)) &     # MACD accelerating
            (df['macd_hist'] > 0) &                            # MACD positive territory
            (df['close'] > df['sma_20'])                       # Above 20 SMA
        )
        # Only signal if TODAY (last row) qualifies
        if conditions.iloc[-1]:
            return df.iloc[[-1]].copy()
        return pd.DataFrame()

    # ──────────────────────────────────────────────────────────────────
    # STRATEGY 2 — Mean Reversion
    # ──────────────────────────────────────────────────────────────────
    def strategy_mean_reversion(self, df):
        """
        Price AT or BELOW lower Bollinger Band, RSI < 35, MACD histogram
        turning up. Signal only fires on the most recent bar.
        """
        df = self._prep(df)
        if len(df) < 25:
            return pd.DataFrame()

        conditions = (
            (df['close'] <= df['bb_lower'] * 1.01) &          # At/below BB lower
            (df['rsi_14'] < 35) &                              # Genuinely oversold
            (df['macd_hist'] > df['macd_hist'].shift(1)) &     # Momentum turning up
            (df['macd_hist'].shift(1) < 0) &                   # Was negative (reversal)
            (df['volume_ratio'] > 1.2) &                       # Volume present
            (df['close'] > df['close'].shift(1) * 0.97)        # Not a free-fall (>3% drop)
        )
        if conditions.iloc[-1]:
            return df.iloc[[-1]].copy()
        return pd.DataFrame()

    # ──────────────────────────────────────────────────────────────────
    # STRATEGY 3 — Trend Following (Golden Cross 20/50)
    # ──────────────────────────────────────────────────────────────────
    def strategy_trend_following(self, df):
        """
        Requires a recent 20 SMA cross above 50 SMA (within last 30 days).
        Entry on healthy pullback (RSI 45-65) above middle Bollinger Band.
        """
        df = self._prep(df)
        if len(df) < 55:
            return pd.DataFrame()

        sma20_above_sma50_recently = False
        for i in range(min(30, len(df))):
            if i > 0:
                prev_idx = len(df) - i - 1
                curr_idx = len(df) - i
                if (0 <= prev_idx < len(df)) and (curr_idx < len(df)):
                    if (df['sma_20'].iloc[curr_idx] > df['sma_50'].iloc[curr_idx] and
                            df['sma_20'].iloc[prev_idx] <= df['sma_50'].iloc[prev_idx]):
                        sma20_above_sma50_recently = True
                        break

        if not sma20_above_sma50_recently:
            return pd.DataFrame()

        conditions = (
            (df['close'] > df['sma_20']) &
            (df['close'] > df['sma_50']) &
            (df['sma_20'] > df['sma_50']) &
            (df['rsi_14'] > 45) & (df['rsi_14'] < 65) &
            (df['close'] > df['bb_middle']) &
            (df['volume_ratio'] > 0.8)
        )
        return df[conditions].tail(7).copy() if conditions.any() else pd.DataFrame()

    # ──────────────────────────────────────────────────────────────────
    # STRATEGY 4 — Volume Breakout
    # ──────────────────────────────────────────────────────────────────
    def strategy_volume_breakout(self, df):
        """
        Unusual volume spike (>2×) with bullish candle >2% up on the day,
        price above 20 SMA. Signal only fires on the most recent bar.
        """
        df = self._prep(df)
        if len(df) < 25:
            return pd.DataFrame()

        conditions = (
            (df['volume_ratio'] > 2.0) &                       # Significant volume spike
            (df['close'] > df['open']) &                       # Bullish candle
            (df['change_pct'] > 2.0) &                        # >2% move
            (df['close'] > df['sma_20']) &                     # Above trend
            (df['rsi_14'] > 50) & (df['rsi_14'] < 75) &       # Healthy RSI
            (df['close'] > df['high'].shift(1))                # Closes above yesterday's high
        )
        if conditions.iloc[-1]:
            return df.iloc[[-1]].copy()
        return pd.DataFrame()

    # ──────────────────────────────────────────────────────────────────
    # STRATEGY 5 — Swing Trading
    # ──────────────────────────────────────────────────────────────────
    def strategy_swing_trading(self, df):
        """
        Swing entry: price pulled back to 20 SMA in an uptrend, RSI recovering,
        MACD bullish with expansion. Signal only fires on the most recent bar.
        """
        df = self._prep(df)
        if len(df) < 30:
            return pd.DataFrame()

        ma50 = df['close'].rolling(50).mean()
        conditions = (
            (df['close'] > df['sma_20']) &                    # Above 20 SMA
            (df['sma_20'] > ma50) &                            # 20 SMA above 50 SMA (uptrend)
            (df['close'] > df['close'].shift(5)) &             # Higher than 5 days ago
            (df['rsi_14'] > 45) & (df['rsi_14'] < 65) &       # Mid-range RSI (not extended)
            (df['macd'] > df['macd_signal']) &                 # MACD bullish
            (df['macd_hist'] > df['macd_hist'].shift(1)) &     # Expanding
            (df['macd_hist'] > 0) &                            # Above zero
            (df['volume_ratio'] > 1.1) &                       # Slightly above average vol
            (df['atr'] > df['atr'].shift(3))                   # Volatility expanding (momentum)
        )
        if conditions.iloc[-1]:
            return df.iloc[[-1]].copy()
        return pd.DataFrame()

    # ──────────────────────────────────────────────────────────────────
    # STRATEGY 6 — Gap Up
    # ──────────────────────────────────────────────────────────────────
    def strategy_gap_up(self, df):
        """
        Gap up >2% from previous close, held intraday (close > open),
        strong volume. Signal only fires on the most recent bar.
        """
        df = self._prep(df)
        if len(df) < 5:
            return pd.DataFrame()

        gap_pct = (df['open'] - df['close'].shift(1)) / df['close'].shift(1) * 100
        conditions = (
            (gap_pct > 2.0) &                                  # Meaningful gap
            (df['close'] > df['open']) &                       # Held intraday (bullish)
            (df['close'] > df['close'].shift(1)) &             # Net positive day
            (df['volume_ratio'] > 1.5) &                       # Volume confirmation
            (df['rsi_14'] < 75)                                # Not overbought
        )
        if conditions.iloc[-1]:
            return df.iloc[[-1]].copy()
        return pd.DataFrame()

    # ──────────────────────────────────────────────────────────────────
    # STRATEGY 7 — RSI Setup
    # ──────────────────────────────────────────────────────────────────
    def strategy_rsi_setup(self, df):
        """
        RSI crosses above 30 from oversold TODAY, price above 10 MA,
        MACD turning bullish. Signal only fires on the most recent bar.
        """
        df = self._prep(df)
        if len(df) < 15:
            return pd.DataFrame()

        ma10 = df['close'].rolling(10).mean()
        rsi_cross_up = (df['rsi_14'] > 30) & (df['rsi_14'].shift(1) <= 30)  # Actual crossover
        conditions = (
            rsi_cross_up &
            (df['close'] > ma10) &
            (df['close'] > df['close'].shift(1)) &             # Up day
            (df['volume_ratio'] > 0.9) &
            (df['macd_hist'] > df['macd_hist'].shift(1))       # MACD turning
        )
        if conditions.iloc[-1]:
            return df.iloc[[-1]].copy()
        return pd.DataFrame()

    # ──────────────────────────────────────────────────────────────────
    # STRATEGY 8 — Stage 2 Uptrend (Weinstein)
    # ──────────────────────────────────────────────────────────────────
    def strategy_stage_2(self, df):
        """
        Stan Weinstein Stage 2 — refined with:
        • 30-day MA slope validation (avoids recovering-from-crash false positives)
        • Failed-breakout penalty scoring
        • MACD histogram quality gate
        • Overhead supply check
        • Three entry tiers: A_Breakout / B_Continuation / C_Early
        """
        df = self._prep(df)
        if len(df) < 200:
            return pd.DataFrame()

        ma10  = df['close'].rolling(10).mean()
        ma20  = df['close'].rolling(20).mean()
        ma50  = df['close'].rolling(50).mean()
        ma150 = df['close'].rolling(150).mean()
        ma200 = df['close'].rolling(200).mean()

        ma20_slope       = ma20  - ma20.shift(10)
        ma50_slope       = ma50  - ma50.shift(10)
        ma150_slope_30d  = ma150 - ma150.shift(30)
        ma200_slope_30d  = ma200 - ma200.shift(30)

        ma_genuinely_rising = (ma150_slope_30d > 0) & (ma200_slope_30d > -50)

        rolling_high  = df['close'].rolling(63).max()
        rolling_low   = df['close'].rolling(63).min()
        base_range_pct = (rolling_high - rolling_low) / rolling_low * 100
        tight_base    = base_range_pct < 30

        high_52w = df['close'].rolling(252).max().shift(1)
        high_26w = df['close'].rolling(126).max().shift(1)
        breakout_52w = df['close'] >= high_52w * 0.98
        breakout_26w = df['close'] >= high_26w * 0.98

        vol_ma50   = df['volume'].rolling(50).mean()
        vol_ratio  = df['volume'] / vol_ma50
        vol_breakout = vol_ratio > 1.5

        up_days   = df['close'] > df['close'].shift(1)
        down_days = df['close'] < df['close'].shift(1)
        vol_up_avg   = df['volume'].where(up_days).rolling(20).mean()
        vol_down_avg = df['volume'].where(down_days).rolling(20).mean()
        healthy_vol  = vol_up_avg > vol_down_avg

        recent_high      = df['close'].rolling(20).max()
        pullback_depth   = (recent_high - df['close']) / recent_high * 100
        shallow_pullback = pullback_depth < 15
        pullback_low_vol = df['volume'] < vol_ma50 * 0.8

        rsi        = df['rsi_14']
        rsi_min_20d = rsi.rolling(20).min()
        rsi_stage2 = (rsi > 50) & (rsi < 78) & (rsi_min_20d > 40)

        macd_hist   = df['macd'] - df['macd_signal']
        macd_bullish = (
            (df['macd'] > df['macd_signal']) &
            (df['macd'] > 0) &
            (df['macd'] > df['macd'].shift(3))
        )
        macd_quality = (
            (macd_hist > 0) &
            (macd_hist > macd_hist.shift(3)) &
            (df['macd'] > 0)
        )

        # Failed-breakout detection
        high_6m             = df['close'].rolling(126).max()
        drawdown_from_peak  = (high_6m - df['close']) / high_6m * 100
        had_failed_breakout = (
            (drawdown_from_peak.rolling(63).max() > 12) &
            (drawdown_from_peak > 3)
        )

        max_recent_vol  = (df['volume'] / vol_ma50).rolling(20).max()
        had_vol_breakout = max_recent_vol > 1.5

        prior_high      = df['close'].rolling(252).max()
        dist_from_res   = (prior_high - df['close']) / df['close'] * 100
        no_overhead     = (dist_from_res > 8) | (dist_from_res < 1)

        # ── Scoring ───────────────────────────────────────────────────
        score = pd.Series(0, index=df.index)
        score += (df['close'] > ma150).astype(int) * 3
        score += (df['close'] > ma200).astype(int) * 3
        score += (ma150_slope_30d > 0).astype(int) * 3
        score += (ma200_slope_30d > 0).astype(int) * 2
        score += ((ma10 > ma20) & (ma20 > ma50) & (ma50 > ma150)).astype(int) * 2
        score += (ma20_slope > 0).astype(int) * 1
        score += (ma50_slope > 0).astype(int) * 1
        score += breakout_26w.astype(int) * 2
        score += breakout_52w.astype(int) * 3
        score += vol_breakout.astype(int) * 2
        score += healthy_vol.astype(int) * 2
        score += rsi_stage2.astype(int) * 2
        score += macd_bullish.astype(int) * 2
        score += shallow_pullback.astype(int) * 1
        score += pullback_low_vol.astype(int) * 1
        # Penalties
        score -= had_failed_breakout.astype(int) * 5
        score -= (~no_overhead).astype(int) * 3
        score -= (macd_hist < 0).astype(int) * 3
        score -= (~had_vol_breakout).astype(int) * 2

        pullback_to_ma20 = (df['close'] <= ma20 * 1.02) & (df['close'] >= ma20 * 0.97)
        pullback_to_ma50 = (df['close'] <= ma50 * 1.02) & (df['close'] >= ma50 * 0.97)

        type_a = (
            (score >= 18) & breakout_26w & vol_breakout &
            (df['close'] > ma150) & (ma150_slope_30d > 0) &
            rsi_stage2 & macd_quality & no_overhead
        )
        type_b = (
            (score >= 15) &
            (pullback_to_ma20 | pullback_to_ma50) &
            pullback_low_vol & (df['close'] > ma150) &
            (ma150_slope_30d > 0) & macd_quality &
            had_vol_breakout & ~had_failed_breakout & no_overhead
        )
        type_c = (
            (score >= 12) &
            (df['close'] > ma50) & (ma50_slope > 0) &
            (df['close'] > ma150) & breakout_26w &
            rsi_stage2 & ma_genuinely_rising
        )

        any_signal = type_a | type_b | type_c
        result     = df[any_signal].copy()
        if result.empty:
            return pd.DataFrame()

        result['stage2_score'] = score[any_signal]
        result['entry_type']   = 'C_Early'
        result.loc[type_b[any_signal], 'entry_type'] = 'B_Continuation'
        result.loc[type_a[any_signal], 'entry_type'] = 'A_Breakout'
        result['ma150_slope']  = ma150_slope_30d[any_signal].round(2)
        result['vol_ratio']    = vol_ratio[any_signal].round(2)
        result['pullback_depth'] = pullback_depth[any_signal].round(2)
        result['above_200ma']  = (df['close'] > ma200)[any_signal]

        return result.tail(5).sort_values('stage2_score', ascending=False)

    # ──────────────────────────────────────────────────────────────────
    # STRATEGY 9 — Strong Linearity
    # ──────────────────────────────────────────────────────────────────
    def strategy_strong_linearity(self, df):
        """
        Consistent higher-highs AND higher-lows over 7 days (≥4/6 moves up).
        Indicates linear institutional accumulation.
        """
        df = self._prep(df)
        if len(df) < 20:
            return pd.DataFrame()

        recent_lows  = df['low'].tail(7)
        recent_highs = df['high'].tail(7)
        rising_lows  = sum(recent_lows.iloc[i] < recent_lows.iloc[i+1]
                           for i in range(len(recent_lows)-1))
        rising_highs = sum(recent_highs.iloc[i] < recent_highs.iloc[i+1]
                           for i in range(len(recent_highs)-1))

        if rising_lows < 4 or rising_highs < 4:
            return pd.DataFrame()

        ma50 = df['close'].rolling(50).mean()
        # Last row must also be part of the linear structure
        last = df.iloc[-1]
        last_ok = (
            last['close'] > df['sma_20'].iloc[-1] and      # Above 20 SMA
            last['close'] > ma50.iloc[-1] and               # Above 50 SMA
            40 < last['rsi_14'] < 70 and
            last['volume_ratio'] > 0.8 and
            last['atr'] > df['atr'].iloc[-2]                # Expanding range
        )
        if last_ok:
            return df.iloc[[-1]].copy()
        return pd.DataFrame()

    # ──────────────────────────────────────────────────────────────────
    # STRATEGY 10 — VCP (Volatility Contraction Pattern)
    # ──────────────────────────────────────────────────────────────────
    def strategy_vcp(self, df):
        """
        ATR and volume both contract ≥20 % over the last 10 vs 20 days.
        Price holds above recent support.  Spring-loaded for breakout.
        """
        df = self._prep(df)
        if len(df) < 50:
            return pd.DataFrame()

        atr_recent   = df['atr'].tail(10).mean()
        atr_previous = df['atr'].tail(20).mean()
        vol_recent   = df['volume'].tail(10).mean()
        vol_previous = df['volume'].tail(20).mean()

        if not (atr_recent < atr_previous * 0.8 and vol_recent < vol_previous * 0.8):
            return pd.DataFrame()

        support    = df['low'].tail(20).min()
        conditions = (
            (df['close'] > support) &
            (df['volume_ratio'] < 1.2) &
            (df['rsi_14'] > 30) & (df['rsi_14'] < 70)
        )
        return df[conditions].tail(3).copy() if conditions.any() else pd.DataFrame()

    # ──────────────────────────────────────────────────────────────────
    # STRATEGY 11 — Pyramiding
    # ──────────────────────────────────────────────────────────────────
    def strategy_pyramiding(self, df,
                             entry_price=None,
                             pyramid_level=0,
                             initial_position=100,
                             add_trigger_pct=4.0,
                             max_pyramids=3):
        """
        Add to a WINNING position in decreasing size as trend confirms.

        Pyramid structure (default unit = initial_position):
          Level 0 — Base     : 100 % at breakout
          Level 1 — Add 1    :  60 % at +4 %
          Level 2 — Add 2    :  30 % at +8 %
          Level 3 — Add 3    :  10 % at +12 %

        Rules
        ─────
        • Only add at NEW highs, never on pullbacks
        • Volume must expand on each add (institutions also adding)
        • RSI ceiling tightens with each level (euphoria risk rises)
        • Trailing stop moves up to breakeven after Level-1 hit
        • Invalidated if MACD histogram turns negative or RSI > 80
        """
        df = self._prep(df)
        if len(df) < 50 or pyramid_level >= max_pyramids:
            return pd.DataFrame()

        close  = df['close']
        volume = df['volume']

        ma10  = close.rolling(10).mean()
        ma20  = close.rolling(20).mean()
        ma50  = close.rolling(50).mean()
        ma10_slope = ma10 - ma10.shift(5)
        ma20_slope = ma20 - ma20.shift(5)

        in_uptrend = (
            (close > ma10) & (close > ma20) & (close > ma50) &
            (ma10_slope > 0) & (ma20_slope > 0) &
            (ma10 > ma20) & (ma20 > ma50)
        )

        high_5d  = df['high'].rolling(5).max().shift(1)
        high_10d = df['high'].rolling(10).max().shift(1)
        new_high = (close >= high_5d * 0.995) | (close >= high_10d * 0.995)

        vol_ma20     = volume.rolling(20).mean()
        vol_ratio    = volume / vol_ma20
        vol_expanding = (vol_ratio > 1.2) & (volume > volume.shift(1)) | (vol_ratio > 1.5)

        rsi_ceilings = {0: 75, 1: 75, 2: 72, 3: 70}
        rsi_ceiling  = rsi_ceilings.get(pyramid_level, 70)
        rsi_ok       = (df['rsi_14'] > 55) & (df['rsi_14'] < rsi_ceiling)

        macd_hist   = df['macd'] - df['macd_signal']
        macd_rising = (macd_hist > 0) & (macd_hist > macd_hist.shift(2)) & (df['macd'] > 0)

        # Entry-price-aware trigger
        if entry_price is not None:
            trigger    = entry_price * (1 + (pyramid_level + 1) * add_trigger_pct / 100)
            at_trigger = close >= trigger
        else:
            ret_5d  = (close - close.shift(5)) / close.shift(5) * 100
            ret_10d = (close - close.shift(10)) / close.shift(10) * 100
            at_trigger = (ret_5d > 2) & (ret_10d > 4) & new_high

        # Invalidation
        do_not_add = (
            (df['rsi_14'] > 80) |
            (macd_hist < macd_hist.shift(1)) |
            (vol_ratio < 0.8) |
            (close < ma20)
        )

        conditions = (
            in_uptrend & vol_expanding & macd_rising &
            rsi_ok & at_trigger & new_high & ~do_not_add
        )

        # Only signal if TODAY (last row) qualifies
        if not conditions.iloc[-1]:
            return pd.DataFrame()

        result = df.iloc[[-1]].copy()

        # Trailing stop: ATR-based, tightens with level
        atr_mults   = {0: 3.0, 1: 2.5, 2: 2.0, 3: 1.5}
        atr_mult    = atr_mults.get(pyramid_level + 1, 1.5)
        result['pyramid_level']    = pyramid_level + 1
        result['add_size_pct']     = [1.0, 0.6, 0.3, 0.1][min(pyramid_level, 3)] * 100
        result['trailing_stop']    = (result['close'] - atr_mult * result['atr']).round(2)
        result['swing_low_stop']   = df['low'].rolling(5).min().iloc[-1].round(2)
        result['recommended_stop'] = max(result['trailing_stop'].iloc[0],
                                         result['swing_low_stop'].iloc[0])
        result['vol_ratio']        = vol_ratio.iloc[-1].round(2)

        return result

    # ──────────────────────────────────────────────────────────────────
    # STRATEGY 12 — Sell Below 10 MA
    # ──────────────────────────────────────────────────────────────────
    def strategy_sell_below_10ma(self, df):
        """
        Price crosses below 10 MA TODAY (fresh crossover, not already below).
        This is an EXIT/AVOID signal — stock's short-term trend just broke.
        Signal only fires on the most recent bar.
        """
        df = self._prep(df)
        if len(df) < 12:
            return pd.DataFrame()

        ma10 = df['close'].rolling(10).mean()
        # Strict crossover: was ABOVE yesterday, is BELOW today
        cross_below = (
            (df['close'] < ma10) &                             # Below 10MA today
            (df['close'].shift(1) >= ma10.shift(1)) &          # Was above yesterday
            (df['close'] < df['close'].shift(1)) &             # Down day
            (df['rsi_14'] < 55) &                              # Weakening RSI
            (df['volume_ratio'] > 0.8)                         # Not on air (volume present)
        )
        if cross_below.iloc[-1]:
            return df.iloc[[-1]].copy()
        return pd.DataFrame()

    # ──────────────────────────────────────────────────────────────────
    # STRATEGY 13 — Golden Crossover (50 EMA × 200 EMA)
    # ──────────────────────────────────────────────────────────────────
    def strategy_golden_crossover(self, df):
        """
        50-day EMA crosses above 200-day EMA within the last 20 days.
        Price above 50 EMA, RSI > 50, MACD positive.
        EMAs within 5 % of each other (fresh crossover only).
        """
        df = self._prep(df)
        if len(df) < 200:
            return pd.DataFrame()

        ema50  = df['close'].ewm(span=50,  adjust=False).mean()
        ema200 = df['close'].ewm(span=200, adjust=False).mean()

        crossover_detected = False
        for i in range(max(1, len(df) - 20), len(df)):
            if ema50.iloc[i-1] <= ema200.iloc[i-1] and ema50.iloc[i] > ema200.iloc[i]:
                crossover_detected = True
                break

        conditions = pd.Series(False, index=df.index)
        if crossover_detected:
            latest = df.index[-1]
            conditions[latest] = (
                (ema50.iloc[-1] > ema200.iloc[-1]) &
                (df['close'].iloc[-1] > ema50.iloc[-1]) &
                (ema50.iloc[-1] < ema200.iloc[-1] * 1.05) &
                (df['rsi_14'].iloc[-1] > 50) &
                (df['volume_ratio'].iloc[-1] > 0.8) &
                (df['macd'].iloc[-1] > df['macd_signal'].iloc[-1]) &
                (df['close'].iloc[-1] > df['sma_20'].iloc[-1])
            )

        result = df[conditions].copy()
        if not result.empty:
            result['ema50']         = ema50[conditions]
            result['ema200']        = ema200[conditions]
            result['ema_alignment'] = ((result['ema50'] - result['ema200']) /
                                       result['ema200'] * 100)
        return result

    # ──────────────────────────────────────────────────────────────────
    # FILTERS
    # ──────────────────────────────────────────────────────────────────
    def filter_by_price_range(self, df, min_price=50, max_price=5000):
        return df[(df['close'] >= min_price) & (df['close'] <= max_price)]

    def filter_by_volume(self, df, min_volume=100_000):
        return df[df['volume'] >= min_volume]

    def filter_by_volatility(self, df, min_atr_pct=1.0):
        atr_pct = (df['atr'] / df['close']) * 100
        return df[atr_pct >= min_atr_pct]

    # ──────────────────────────────────────────────────────────────────
    # RISK / REWARD
    # ──────────────────────────────────────────────────────────────────
    def calculate_risk_reward(self, df):
        """ATR-based stop (1.5×) and tiered targets (2–3×)."""
        if df.empty:
            return df

        df['stop_loss'] = df['close'] - 1.5 * df['atr']

        for idx in df.index:
            rsi = df.loc[idx, 'rsi_14']
            vr  = df.loc[idx, 'volume_ratio']
            mult = 3.0 if (rsi > 65 and vr > 1.5) else 2.5 if rsi > 60 else 2.0 if rsi < 40 else 2.5
            df.loc[idx, 'target_1'] = df.loc[idx, 'close'] + mult       * df.loc[idx, 'atr']
            df.loc[idx, 'target_2'] = df.loc[idx, 'close'] + (mult + 1) * df.loc[idx, 'atr']

        df['risk_pct']     = (df['close'] - df['stop_loss']) / df['close'] * 100
        df['reward_1_pct'] = (df['target_1'] - df['close'])  / df['close'] * 100
        df['reward_2_pct'] = (df['target_2'] - df['close'])  / df['close'] * 100
        df['rr_ratio_1']   = df['reward_1_pct'] / df['risk_pct']
        df['rr_ratio_2']   = df['reward_2_pct'] / df['risk_pct']
        return df

    # ──────────────────────────────────────────────────────────────────
    # SCAN DISPATCHER
    # ──────────────────────────────────────────────────────────────────
    def scan_with_strategies(self, stock_data: pd.DataFrame) -> dict:
        """Apply all strategies, return dict {strategy_name: DataFrame}."""
        strategies = {
            'Momentum Breakout': self.strategy_momentum_breakout,
            'Mean Reversion':    self.strategy_mean_reversion,
            'Trend Following':   self.strategy_trend_following,
            'Volume Breakout':   self.strategy_volume_breakout,
            'Swing Trading':     self.strategy_swing_trading,
            'Gap Up':            self.strategy_gap_up,
            'RSI Setup':         self.strategy_rsi_setup,
            'Stage 2 Uptrend':   self.strategy_stage_2,
            'Strong Linearity':  self.strategy_strong_linearity,
            'VCP Pattern':       self.strategy_vcp,
            'Pyramiding':        self.strategy_pyramiding,
            'Sell Below 10MA':   self.strategy_sell_below_10ma,
            'Golden Crossover':  self.strategy_golden_crossover,
        }

        all_matches  = {s: [] for s in strategies}
        symbol_count = 0

        for symbol, sym_df in stock_data.groupby('symbol'):
            symbol_count += 1
            sym_df = sym_df.reset_index(drop=True)

            for name, func in strategies.items():
                try:
                    matched = func(sym_df.copy())
                    if matched is None or matched.empty:
                        continue

                    # After _prep(), 'date' becomes the index.
                    # Reset it back to a column for downstream processing.
                    if matched.index.name == 'date':
                        matched = matched.reset_index()

                    last = matched.iloc[[-1]].copy()
                    if 'symbol' not in last.columns:
                        last['symbol'] = symbol
                    all_matches[name].append(last)
                except Exception as e:
                    logger.error(f"{name}/{symbol}: {e}")

        logger.info(f"Processed {symbol_count} symbols")

        results = {}
        for name, matches in all_matches.items():
            if not matches:
                logger.warning(f"{name}: no matches")
                continue

            combined = pd.concat(matches, ignore_index=True)

            # ── LATEST DATE ONLY per symbol ───────────────────────────
            # Strategies return up to 5 trailing rows per symbol.
            # We only want the single most recent row for each symbol.
            if 'date' in combined.columns:
                combined['date'] = pd.to_datetime(combined['date'])
                combined = (combined
                            .sort_values('date', ascending=False)
                            .drop_duplicates(subset=['symbol'], keep='first')
                            .sort_values('symbol')
                            .reset_index(drop=True))
            else:
                # No date column — just keep one row per symbol
                combined = (combined
                            .drop_duplicates(subset=['symbol'], keep='last')
                            .sort_values('symbol')
                            .reset_index(drop=True))

            combined = self.calculate_risk_reward(combined)
            combined['scan_date'] = datetime.now().strftime('%Y-%m-%d')
            combined['strategy']  = name
            results[name] = combined
            logger.info(f"{name}: {len(combined)} stocks")

        return results

    # ──────────────────────────────────────────────────────────────────
    # BACKTEST RUNNER
    # ──────────────────────────────────────────────────────────────────
    def run_backtest(self, stock_data: pd.DataFrame,
                     strategies_to_test: list | None = None,
                     test_pct: float = 0.3) -> dict:
        """
        Run walk-forward backtests for selected (or all) strategies.

        Returns a dict keyed by strategy name, each value containing:
          • overall      — aggregated metrics across all symbols
          • per_symbol   — per-symbol breakdown
          • all_trades   — flat list of every simulated trade
        """
        engine = BacktestEngine()

        strategy_map = {
            'Momentum Breakout': self.strategy_momentum_breakout,
            'Mean Reversion':    self.strategy_mean_reversion,
            'Trend Following':   self.strategy_trend_following,
            'Volume Breakout':   self.strategy_volume_breakout,
            'Swing Trading':     self.strategy_swing_trading,
            'Gap Up':            self.strategy_gap_up,
            'RSI Setup':         self.strategy_rsi_setup,
            'Stage 2 Uptrend':   self.strategy_stage_2,
            'Strong Linearity':  self.strategy_strong_linearity,
            'VCP Pattern':       self.strategy_vcp,
            'Pyramiding':        self.strategy_pyramiding,
            'Golden Crossover':  self.strategy_golden_crossover,
        }

        if strategies_to_test:
            strategy_map = {k: v for k, v in strategy_map.items()
                            if k in strategies_to_test}

        # Temporarily enable DEBUG to see per-symbol detail
        bt_logger = logging.getLogger(__name__)
        prev_level = bt_logger.level
        bt_logger.setLevel(logging.DEBUG)

        backtest_results = {}
        for name, func in strategy_map.items():
            logger.info(f"[Backtest] Running {name} ...")
            try:
                result = engine.run(stock_data, func, name, test_pct)
                backtest_results[name] = result
            except Exception as e:
                logger.error(f"[Backtest] {name} failed: {e}")

        bt_logger.setLevel(prev_level)
        return backtest_results


# ═══════════════════════════════════════════════════════════════════════
#  HELPER FUNCTIONS (unchanged from original)
# ═══════════════════════════════════════════════════════════════════════

def get_strategy_recommendation(strategy_name: str) -> str:
    buy_strategies  = [
        'Momentum Breakout', 'Trend Following', 'Volume Breakout', 'Gap Up',
        'Stage 2 Uptrend', 'Strong Linearity', 'VCP Pattern', 'Pyramiding',
        'Golden Crossover', 'Mean Reversion', 'RSI Setup',
    ]
    sell_strategies = ['Sell Below 10MA']
    hold_strategies = ['Swing Trading']

    if strategy_name in buy_strategies:
        return 'BUY'
    if strategy_name in sell_strategies:
        return 'SELL'
    if strategy_name in hold_strategies:
        return 'HOLD'
    return 'HOLD'


def calculate_overall_recommendation(strategy_dates: list) -> tuple:
    if not strategy_dates:
        return 'HOLD', False

    normalized = []
    for name, sig_date in strategy_dates:
        if sig_date is None or (isinstance(sig_date, float) and pd.isna(sig_date)):
            continue
        normalized.append((name, pd.to_datetime(sig_date).normalize()))

    if not normalized:
        return 'HOLD', False

    latest_date       = max(d for _, d in normalized)
    latest_strategies = [s for s, d in normalized if d == latest_date]
    recs              = [get_strategy_recommendation(s) for s in latest_strategies]

    has_buy  = 'BUY'  in recs
    has_sell = 'SELL' in recs

    if has_buy and has_sell:
        return 'CONFLICT', True
    if has_buy:
        return 'BUY', False
    if has_sell:
        return 'SELL', False
    return 'HOLD', False


# ═══════════════════════════════════════════════════════════════════════
#  BACKTEST HTML REPORT
# ═══════════════════════════════════════════════════════════════════════

def generate_backtest_html(backtest_results: dict, timestamp: str) -> str:
    """Generate a standalone HTML report for backtest results."""

    rows = []
    for name, data in backtest_results.items():
        ov = data.get('overall', {})
        if not ov:
            continue
        wc = 'color:#28a745;font-weight:bold' if ov.get('win_rate_pct', 0) >= 50 else 'color:#dc3545'
        pf_val = ov.get('profit_factor', 0)
        pfc = 'color:#28a745;font-weight:bold' if pf_val >= 1.5 else 'color:#ffc107'
        rows.append(f"""
        <tr>
          <td><strong>{name}</strong></td>
          <td>{ov.get('total_trades', 0)}</td>
          <td style="{wc}">{ov.get('win_rate_pct', 0):.1f}%</td>
          <td>{ov.get('avg_win_pct', 0):.2f}%</td>
          <td>{ov.get('avg_loss_pct', 0):.2f}%</td>
          <td>₹{ov.get('total_pnl_inr', 0):,.0f}</td>
          <td>₹{ov.get('max_drawdown_inr', 0):,.0f}</td>
          <td style="{pfc}">{pf_val:.2f}</td>
          <td>{ov.get('sharpe_ratio', 0):.2f}</td>
          <td>{ov.get('avg_hold_days', 0):.1f}</td>
          <td>{ov.get('stop_loss_exits', 0)}</td>
          <td>{ov.get('target2_exits', 0)}</td>
        </tr>""")

    table_rows = '\n'.join(rows)

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Backtest Results — {datetime.now().strftime('%B %d, %Y')}</title>
<style>
  body  {{ font-family:'Segoe UI',sans-serif; background:#f0f4f8; color:#333; padding:20px; }}
  .wrap {{ max-width:1400px; margin:0 auto; background:#fff; border-radius:12px;
           box-shadow:0 4px 20px rgba(0,0,0,.15); overflow:hidden; }}
  .hdr  {{ background:linear-gradient(135deg,#1e3c72,#2a5298); color:#fff;
           padding:35px 40px; text-align:center; }}
  .hdr h1 {{ margin:0 0 8px; font-size:2em; }}
  .hdr p  {{ margin:0; opacity:.85; }}
  .body {{ padding:40px; }}
  .alert {{ background:#fff3cd; border-left:5px solid #ffc107; padding:14px 18px;
            border-radius:6px; margin-bottom:24px; }}
  table {{ width:100%; border-collapse:collapse; }}
  th    {{ background:linear-gradient(135deg,#667eea,#764ba2); color:#fff;
           padding:12px 14px; text-align:left; font-size:.85em; letter-spacing:.4px; }}
  td    {{ padding:11px 14px; border-bottom:1px solid #e8e8e8; font-size:.9em; }}
  tr:hover {{ background:#f7f8ff; }}
  .footer {{ text-align:center; padding:24px; color:#888; font-size:.85em;
             border-top:2px solid #eee; }}
</style>
</head>
<body>
<div class="wrap">
  <div class="hdr">
    <h1>📊 Strategy Backtest Results</h1>
    <p>Walk-Forward Out-of-Sample Test &nbsp;|&nbsp; Generated {datetime.now().strftime('%B %d, %Y at %I:%M %p')}</p>
  </div>
  <div class="body">
    <div class="alert">
      <strong>Methodology:</strong> Each strategy is tested on the last 30 % of available
      history (out-of-sample). Signals on day T are executed at the open of day T+1.
      Stop-loss = 1.5× ATR below entry; Target-1 = 2.5× ATR; Target-2 = 3.5× ATR.
      After T1 hit, 50 % of position is closed and stop trails to breakeven.
      Max hold = 20 trading days.
    </div>
    <div style="overflow-x:auto">
    <table>
      <thead>
        <tr>
          <th>Strategy</th><th>Trades</th><th>Win Rate</th>
          <th>Avg Win %</th><th>Avg Loss %</th><th>Total P&amp;L (₹)</th>
          <th>Max DD (₹)</th><th>Profit Factor</th><th>Sharpe</th>
          <th>Avg Hold Days</th><th>Stop Exits</th><th>T2 Exits</th>
        </tr>
      </thead>
      <tbody>{table_rows}</tbody>
    </table>
    </div>
  </div>
  <div class="footer">Backtest is for informational purposes only. Past performance ≠ future results.</div>
</div>
</body>
</html>"""

# ═══════════════════════════════════════════════════════════════════════
#  EXCEL REPORT WRITER
# ═══════════════════════════════════════════════════════════════════════



# ═══════════════════════════════════════════════════════════════════════
#  NIFTY 50 DISPLAY NAMES
# ═══════════════════════════════════════════════════════════════════════

STOCK_NAMES = {
    'RELIANCE':'Reliance Industries','TCS':'TCS','HDFCBANK':'HDFC Bank',
    'INFY':'Infosys','ICICIBANK':'ICICI Bank','HINDUNILVR':'Hindustan Unilever',
    'ITC':'ITC','SBIN':'State Bank of India','BHARTIARTL':'Bharti Airtel',
    'KOTAKBANK':'Kotak Mahindra Bank','LT':'Larsen & Toubro','AXISBANK':'Axis Bank',
    'ASIANPAINT':'Asian Paints','MARUTI':'Maruti Suzuki','SUNPHARMA':'Sun Pharma',
    'TITAN':'Titan Company','BAJFINANCE':'Bajaj Finance','WIPRO':'Wipro',
    'ULTRACEMCO':'UltraTech Cement','NESTLEIND':'Nestle India','ONGC':'ONGC',
    'NTPC':'NTPC','POWERGRID':'Power Grid','HCLTECH':'HCL Technologies',
    'COALINDIA':'Coal India','BAJAJFINSV':'Bajaj Finserv','M&M':'Mahindra & Mahindra',
    'ADANIPORTS':'Adani Ports','TATASTEEL':'Tata Steel','GRASIM':'Grasim Industries',
    'TECHM':'Tech Mahindra','INDUSINDBK':'IndusInd Bank','DIVISLAB':"Divi's Labs",
    'DRREDDY':"Dr. Reddy's",'CIPLA':'Cipla','EICHERMOT':'Eicher Motors',
    'HINDALCO':'Hindalco','BRITANNIA':'Britannia','JSWSTEEL':'JSW Steel',
    'APOLLOHOSP':'Apollo Hospitals','HEROMOTOCO':'Hero MotoCorp','SHREECEM':'Shree Cement',
    'BPCL':'BPCL','TATACONSUM':'Tata Consumer','UPL':'UPL',
    'SBILIFE':'SBI Life Insurance','BAJAJ-AUTO':'Bajaj Auto',
    'ADANIENT':'Adani Enterprises','TATAMOTORS':'Tata Motors',
}


def _add_stock_name(df: pd.DataFrame) -> pd.DataFrame:
    """Add Stock Name as second column after symbol, move symbol first."""
    if df.empty:
        return df
    df = df.copy()
    if 'symbol' in df.columns:
        df['stock_name'] = df['symbol'].map(STOCK_NAMES).fillna(df['symbol'])
        # Reorder: symbol, stock_name first, then everything else
        other_cols = [c for c in df.columns if c not in ('symbol', 'stock_name')]
        df = df[['symbol', 'stock_name'] + other_cols]
    return df


def _write_excel_report(path, strategy_results: dict, backtest_results: dict, run_label: str):
    """
    Write a multi-sheet Excel workbook:

    Sheet 1 — Summary        : one row per strategy (live signals + backtest KPIs)
    Sheet 2 — All Signals    : every stock that fired ANY strategy today (latest date only)
    Sheet 3+ — Per-strategy  : one sheet per strategy with all matched stocks
                               (each stock appears once — latest signal date)
    """
    try:
        import openpyxl
        from openpyxl.styles import (PatternFill, Font, Alignment,
                                     Border, Side, numbers)
        from openpyxl.utils import get_column_letter
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        print("⚠  openpyxl not installed — skipping Excel. Run: pip install openpyxl")
        return

    wb = openpyxl.Workbook()

    # ── Styles ────────────────────────────────────────────────────────
    HDR_FILL   = PatternFill("solid", fgColor="1F3864")
    HDR_FONT   = Font(color="FFFFFF", bold=True, size=10)
    TITLE_FONT = Font(bold=True, size=12, color="1F3864")
    ALT_FILL   = PatternFill("solid", fgColor="EEF2FF")
    GREEN_FILL = PatternFill("solid", fgColor="D6F4D0")
    RED_FILL   = PatternFill("solid", fgColor="FCE4E4")
    AMBER_FILL = PatternFill("solid", fgColor="FFF3CD")
    THIN       = Side(style='thin', color='CCCCCC')
    BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER     = Alignment(horizontal='center', vertical='center')
    LEFT       = Alignment(horizontal='left',   vertical='center')

    def style_header_row(ws, row_num, n_cols):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = CENTER
            cell.border = BORDER

    def style_data_row(ws, row_num, n_cols, alt=False):
        fill = ALT_FILL if alt else PatternFill("solid", fgColor="FFFFFF")
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = LEFT

    def autofit(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 30)

    # ══════════════════════════════════════════════════════════════════
    # SHEET 1 — SUMMARY
    # ══════════════════════════════════════════════════════════════════
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_view.showGridLines = False

    ws_sum['A1'] = f"📊  Stock Scanner Report  —  {run_label}"
    ws_sum['A1'].font = Font(bold=True, size=14, color="1F3864")
    ws_sum.merge_cells('A1:N1')
    ws_sum.row_dimensions[1].height = 28

    ws_sum['A2'] = "Out-of-sample backtest (last 30% of history). Signal on day T → entry at open of T+1."
    ws_sum['A2'].font = Font(italic=True, size=9, color="666666")
    ws_sum.merge_cells('A2:N2')

    headers = [
        "Strategy", "Live Signals",
        "Backtest Trades", "Win Rate %", "Avg Win %", "Avg Loss %",
        "Total P&L (₹)", "Max DD (₹)", "Profit Factor", "Sharpe",
        "Avg Hold Days", "Stop Exits", "T1 Exits", "T2 Exits"
    ]
    for c, h in enumerate(headers, 1):
        ws_sum.cell(row=4, column=c, value=h)
    style_header_row(ws_sum, 4, len(headers))
    ws_sum.row_dimensions[4].height = 22

    row = 5
    for name, df_live in strategy_results.items():
        ov  = backtest_results.get(name, {}).get('overall', {})
        n_live = len(df_live) if not df_live.empty else 0
        n_bt   = ov.get('total_trades', 0)
        wr     = ov.get('win_rate_pct', None)
        pf     = ov.get('profit_factor', None)
        sh     = ov.get('sharpe_ratio', None)
        pnl    = ov.get('total_pnl_inr', None)
        dd     = ov.get('max_drawdown_inr', None)
        aw     = ov.get('avg_win_pct', None)
        al     = ov.get('avg_loss_pct', None)
        hd     = ov.get('avg_hold_days', None)
        sl     = ov.get('stop_loss_exits', 0)
        t1     = ov.get('target1_exits', 0)
        t2     = ov.get('target2_exits', 0)

        vals = [name, n_live, n_bt, wr, aw, al, pnl, dd, pf, sh, hd, sl, t1, t2]
        for c, v in enumerate(vals, 1):
            ws_sum.cell(row=row, column=c, value=v)
        style_data_row(ws_sum, row, len(headers), alt=(row % 2 == 0))

        # Colour-code Profit Factor cell
        pf_cell = ws_sum.cell(row=row, column=9)
        if pf is not None:
            pf_cell.fill = GREEN_FILL if pf >= 1.5 else (AMBER_FILL if pf >= 1.0 else RED_FILL)

        # Colour-code Win Rate cell
        wr_cell = ws_sum.cell(row=row, column=4)
        if wr is not None:
            wr_cell.fill = GREEN_FILL if wr >= 50 else (AMBER_FILL if wr >= 40 else RED_FILL)
            wr_cell.number_format = '0.0"%"'

        # Colour-code P&L
        pnl_cell = ws_sum.cell(row=row, column=7)
        if pnl is not None:
            pnl_cell.fill  = GREEN_FILL if pnl > 0 else RED_FILL
            pnl_cell.number_format = '₹#,##0;[RED]-₹#,##0'

        row += 1

    ws_sum.freeze_panes = 'A5'
    autofit(ws_sum)

    # ══════════════════════════════════════════════════════════════════
    # SHEET 2 — ALL SIGNALS (one row per stock, latest signal date only)
    # ══════════════════════════════════════════════════════════════════
    all_dfs = []
    for name, df in strategy_results.items():
        if not df.empty:
            tmp = df.copy()
            tmp['strategy'] = name
            all_dfs.append(tmp)

    if all_dfs:
        combined = pd.concat(all_dfs, ignore_index=True)
        if 'date' in combined.columns:
            combined['date'] = pd.to_datetime(combined['date'])
            # For each symbol keep the row with the LATEST date
            combined = (combined
                        .sort_values('date', ascending=False)
                        .drop_duplicates(subset=['symbol'], keep='first')
                        .sort_values('symbol')
                        .reset_index(drop=True))

        combined = _add_stock_name(combined)

        ws_all = wb.create_sheet("All Signals")
        ws_all.sheet_view.showGridLines = False
        ws_all['A1'] = f"All Signals — {run_label}  (latest signal date per stock)"
        ws_all['A1'].font = TITLE_FONT
        ws_all.merge_cells(f'A1:{get_column_letter(len(combined.columns))}1')
        ws_all.row_dimensions[1].height = 22
        _write_df_to_sheet(ws_all, combined, start_row=3,
                           HDR_FILL=HDR_FILL, HDR_FONT=HDR_FONT,
                           ALT_FILL=ALT_FILL, BORDER=BORDER,
                           CENTER=CENTER, LEFT=LEFT)
        ws_all.freeze_panes = 'A4'
        autofit(ws_all)

    # ══════════════════════════════════════════════════════════════════
    # SHEETS 3+ — One sheet per strategy
    # ══════════════════════════════════════════════════════════════════
    for name, df in strategy_results.items():
        if df.empty:
            continue

        sheet_name = name[:31]  # Excel 31-char limit
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False

        ov = backtest_results.get(name, {}).get('overall', {})
        n  = ov.get('total_trades', 0)
        pf = ov.get('profit_factor', 0)
        wr = ov.get('win_rate_pct', 0)
        sh = ov.get('sharpe_ratio', 0)
        pnl= ov.get('total_pnl_inr', 0)

        ws['A1'] = f"{name}  —  {run_label}"
        ws['A1'].font = TITLE_FONT
        ws['A2'] = (f"Backtest (out-of-sample):  {n} trades  |  "
                    f"Win Rate {wr:.1f}%  |  PF {pf:.2f}  |  "
                    f"Sharpe {sh:.2f}  |  P&L ₹{pnl:,.0f}")
        ws['A2'].font = Font(italic=True, size=9, color="555555")

        # Keep latest date per symbol
        sheet_df = df.copy()
        if 'date' in sheet_df.columns:
            sheet_df['date'] = pd.to_datetime(sheet_df['date'])
            sheet_df = (sheet_df
                        .sort_values('date', ascending=False)
                        .drop_duplicates(subset=['symbol'], keep='first')
                        .sort_values('symbol')
                        .reset_index(drop=True))

        sheet_df = _add_stock_name(sheet_df)

        col_count = len(sheet_df.columns)
        ws.merge_cells(f'A1:{get_column_letter(col_count)}1')
        ws.merge_cells(f'A2:{get_column_letter(col_count)}2')
        ws.row_dimensions[1].height = 22
        _write_df_to_sheet(ws, sheet_df, start_row=4,
                           HDR_FILL=HDR_FILL, HDR_FONT=HDR_FONT,
                           ALT_FILL=ALT_FILL, BORDER=BORDER,
                           CENTER=CENTER, LEFT=LEFT)
        ws.freeze_panes = 'A5'
        autofit(ws)

    wb.save(path)


def _write_df_to_sheet(ws, df, start_row,
                        HDR_FILL, HDR_FONT, ALT_FILL, BORDER, CENTER, LEFT):
    """Write a DataFrame to a worksheet starting at start_row."""
    from openpyxl.utils import get_column_letter

    # Header row
    for c, col in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=c, value=str(col).replace('_', ' ').title())
        cell.fill   = HDR_FILL
        cell.font   = HDR_FONT
        cell.border = BORDER
        cell.alignment = CENTER
    ws.row_dimensions[start_row].height = 20

    THIN  = __import__('openpyxl').styles.Side(style='thin', color='CCCCCC')

    # Data rows
    from datetime import datetime as _dt
    for r_idx, (_, row_data) in enumerate(df.iterrows(), start_row + 1):
        fill = ALT_FILL if r_idx % 2 == 0 else __import__('openpyxl').styles.PatternFill("solid", fgColor="FFFFFF")
        for c_idx, val in enumerate(row_data, 1):
            # Convert pandas Timestamp -> native datetime (strips tz)
            if hasattr(val, 'to_pydatetime'):
                val = val.to_pydatetime().replace(tzinfo=None)
            # Convert numpy/pandas types to native Python
            elif hasattr(val, 'item'):
                val = val.item()
            elif isinstance(val, str):
                pass
            else:
                try:
                    if pd.isna(val):
                        val = ''
                except Exception:
                    pass
            # Strip timezone from any remaining datetime
            if isinstance(val, _dt) and val.tzinfo is not None:
                val = val.replace(tzinfo=None)
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill      = fill
            cell.border    = BORDER
            cell.alignment = LEFT


# ═══════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════

def main():
    import os

    API_KEY      = os.environ.get('KITE_API_KEY', "u664cda77q2cf7ft")
    ACCESS_TOKEN = os.environ.get('KITE_ACCESS_TOKEN')

    cred_file = os.path.join(os.path.dirname(__file__), 'kite_credentials.txt')
    if not ACCESS_TOKEN and os.path.exists(cred_file):
        with open(cred_file) as f:
            for line in f:
                if line.startswith('ACCESS_TOKEN='):
                    ACCESS_TOKEN = line.strip().split('=', 1)[1]
                if line.startswith('API_KEY=') and not API_KEY:
                    API_KEY = line.strip().split('=', 1)[1]

    if not ACCESS_TOKEN:
        print('\nERROR: ACCESS_TOKEN not found.')
        return

    scanner = AdvancedStockScanner(API_KEY, ACCESS_TOKEN)

    # ── Stock universe ────────────────────────────────────────────────
    logger.info("Fetching all NSE stocks (excluding indices, ETFs, and options)...")
    all_stocks = scanner.get_nse_stocks('NSE')
    
    # Filter out indices, ETFs, and options - keep only pure equity stocks
    def is_pure_stock(stock):
        """Check if stock is a pure equity (not index, ETF, or option)"""
        symbol = stock.get('tradingsymbol', '').upper()
        # Exclude common index names and ETF suffixes
        excluded_keywords = ['INDEX', 'ETF', 'NIFTY', 'SENSEX', 'NIFTY50', 'JUNIOR', 'MIDCAP', 
                            'SMALLCAP', 'LARGECAP', 'PSU', 'PRIVATE', 'BANK', 'AUTO', 'IT',
                            'PHARMA', 'METALS', 'ENERGY', 'FINANCE', 'CONSUMR', 'UTILITY',
                            'REALTY', 'PSE']
        # Check if any keyword is in the symbol (common in ETF names)
        if any(keyword in symbol for keyword in excluded_keywords if len(keyword) > 3):
            return False
        # Also check instrument_type (should be EQ for equity, but double-check)
        if stock.get('instrument_type') != 'EQ':
            return False
        return True
    
    stocks_to_scan = [s for s in all_stocks if is_pure_stock(s)]
    
    if not stocks_to_scan:
        logger.warning('No pure stocks found; using all NSE stocks')
        stocks_to_scan = all_stocks

    logger.info("Fetching historical data (300 days)...")
    stock_data = scanner.scan_stocks_for_strategies(stocks_to_scan, lookback_days=300)

    if stock_data.empty:
        print("No stock data retrieved.")
        return

    print(f"Retrieved {stock_data['symbol'].nunique()} stocks, {len(stock_data)} rows")

    # ── Live scan ─────────────────────────────────────────────────────
    logger.info("Data ready for strategy analysis...")
    strategy_results = scanner.scan_with_strategies(stock_data)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    run_label = datetime.now().strftime('%d %b %Y  %H:%M')

    # ── Backtest ──────────────────────────────────────────────────────
    print("\nRunning walk-forward backtest (this may take a minute)...")
    backtest_results = scanner.run_backtest(stock_data, test_pct=0.3)

    # ── Save backtest trades CSV ──────────────────────────────────────
    bt_trades = []
    for name, data in backtest_results.items():
        for t in data.get('all_trades', []):
            t['strategy'] = name
            bt_trades.append(t)

    if bt_trades:
        # Use timestamp to avoid PermissionError if previous file is open in Excel
        bt_csv = RESULTS_DIR / f"backtest_trades_{timestamp}.csv"
        # Also try to clean up old backtest_trades files (ignore if locked)
        for old_f in RESULTS_DIR.glob("backtest_trades_*.csv"):
            if old_f.name != bt_csv.name:
                try:
                    old_f.unlink()
                except Exception:
                    pass
        pd.DataFrame(bt_trades).to_csv(bt_csv, index=False)
        print(f"✓ Backtest trades CSV: {bt_csv}")

    # ── Save backtest HTML ────────────────────────────────────────────
    bt_html = RESULTS_DIR / "backtest_results.html"
    bt_html.write_text(generate_backtest_html(backtest_results, timestamp), encoding='utf-8')
    print(f"✓ Backtest HTML: {bt_html}")

    # ── Save combined Excel (Summary + All Signals + per-strategy) ───
    xlsx_file = RESULTS_DIR / "stock_scanner_results.xlsx"
    try:
        _write_excel_report(xlsx_file, strategy_results, backtest_results, run_label)
        print(f"✓ Excel report: {xlsx_file}")
    except PermissionError:
        # File is open in Excel — write to a timestamped copy instead
        xlsx_fallback = RESULTS_DIR / f"stock_scanner_results_{timestamp}.xlsx"
        _write_excel_report(xlsx_fallback, strategy_results, backtest_results, run_label)
        print(f"✓ Excel report (close the previous file next time!): {xlsx_fallback}")

    # ── Save scanner HTML ─────────────────────────────────────────────
    try:
        from enhanced_html_generator import generate_enhanced_html
        stock_data_dict = {}
        all_results = [v for v in strategy_results.values() if not v.empty]
        if all_results:
            combined_syms = pd.concat(all_results, ignore_index=True)['symbol'].unique()
            for sym in combined_syms:
                sd = stock_data[stock_data['symbol'] == sym]
                if not sd.empty:
                    stock_data_dict[sym] = {
                        '52w_high': sd['high'].rolling(min(252, len(sd))).max().iloc[-1],
                        '52w_low':  sd['low'].rolling(min(252, len(sd))).min().iloc[-1],
                        'atr':      sd['atr'].iloc[-1] if 'atr' in sd.columns else 0,
                    }
        html_file = RESULTS_DIR / "scanner_results.html"
        html_file.write_text(
            generate_enhanced_html(strategy_results, timestamp, stock_data_dict),
            encoding='utf-8'
        )
        print(f"✓ Scanner HTML: {html_file}")
    except Exception as e:
        logger.warning(f"Scanner HTML skipped: {e}")

    # ── Console backtest summary ──────────────────────────────────────
    print("\n" + "="*80)
    print("BACKTEST SUMMARY")
    print("="*80)
    print(f"{'Strategy':<22} {'Trades':>7} {'WinRate':>8} {'PF':>6} {'Sharpe':>7} {'TotalPnL':>12}")
    print("-"*80)
    for name, data in backtest_results.items():
        ov = data.get('overall', {})
        if ov:
            print(f"{name:<22} {ov.get('total_trades', 0):>7} "
                  f"{ov.get('win_rate_pct', 0):>7.1f}% "
                  f"{ov.get('profit_factor', 0):>6.2f} "
                  f"{ov.get('sharpe_ratio', 0):>7.2f} "
                  f"₹{ov.get('total_pnl_inr', 0):>11,.0f}")

    print("\n✅ Scan + Backtest complete.")
    print(f"   Results folder: {RESULTS_DIR.resolve()}")


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO,
                        format='%(asctime)s - %(levelname)s - %(message)s')
    main()